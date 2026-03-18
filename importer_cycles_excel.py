# =============================================================================
# importer_cycles_excel.py
# =============================================================================
# Lit le fichier Excel (Activités_v2_ajouts.xlsx) et remplit la table
# activite_cycle_analysee pour les cycles C1, C2, C3 :
#   - Activite marquee "X" pour un cycle -> non touchee (Gemini la traitera)
#   - Activite NON marquee pour un cycle  -> inseree avec statut='inadapte'
#
# Prerequis : migration_ajout_statut_cycle.py execute (colonne statut presente)
# SIMULATION = True par defaut.
# =============================================================================

import sqlite3
import sys
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl

DB_PATH    = Path(__file__).parent / "activites.db"
EXCEL_PATH = Path(r"C:\Users\moina\Dropbox\Animation\Activités v2\Activités v2 ajouts.xlsx")

SIMULATION = False

# Seuil de correspondance fuzzy (0.0 - 1.0)
# En dessous de ce seuil, l'activite est ignoree avec un avertissement
SEUIL_FUZZY = 0.85


def ratio(a, b):
    return SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio()


def trouver_correspondance(nom_excel, activites_db):
    meilleur = None
    meilleur_score = 0.0
    for act in activites_db:
        s = ratio(nom_excel, act["nom"])
        if s > meilleur_score:
            meilleur_score = s
            meilleur = act
    return meilleur, meilleur_score


def main():
    print("=" * 60)
    print("IMPORT CYCLES DEPUIS EXCEL")
    print("=" * 60)
    print(f"MODE : {'SIMULATION' if SIMULATION else 'ECRITURE EN BASE'}")
    print()

    if not DB_PATH.exists():
        print(f"ERREUR : Base SQLite introuvable : {DB_PATH}")
        sys.exit(1)

    if not EXCEL_PATH.exists():
        print(f"ERREUR : Fichier Excel introuvable : {EXCEL_PATH}")
        sys.exit(1)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    # Verifier que la colonne statut existe
    cols = [r[1] for r in conn.execute("PRAGMA table_info(activite_cycle_analysee)").fetchall()]
    if "statut" not in cols:
        print("ERREUR : colonne 'statut' absente. Lancez migration_ajout_statut_cycle.py")
        conn.close()
        sys.exit(1)

    # Charger les activites en base
    activites_db = [dict(r) for r in conn.execute(
        "SELECT id, nom FROM activite ORDER BY nom"
    ).fetchall()]
    print(f"  {len(activites_db)} activites en base")

    # Charger les cycles
    cycles = {r["code"]: r["id"] for r in conn.execute(
        "SELECT id, code FROM cycle WHERE code IN ('C1','C2','C3')"
    ).fetchall()}
    print(f"  Cycles charges : {list(cycles.keys())}")

    # Lire l'Excel
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Modifications"]
    headers = [c.value for c in ws[1]]
    idx = {
        "nom":  headers.index("Activité"),
        "C1":   headers.index("C1"),
        "C2":   headers.index("C2"),
        "C3":   headers.index("C3"),
    }
    print(f"  {ws.max_row - 1} lignes dans l'Excel")
    print()

    non_trouves   = []
    inadaptes_new = 0   # nouvelles lignes inadapte a inserer
    deja_presents = 0   # lignes deja dans activite_cycle_analysee (non touchees)
    scores_bas    = []

    # Precharger les associations existantes pour eviter les doublons
    deja = {(r[0], r[1]) for r in conn.execute(
        "SELECT activite_id, cycle_id FROM activite_cycle_analysee"
    ).fetchall()}

    for row in ws.iter_rows(min_row=2, values_only=True):
        nom_excel = row[idx["nom"]]
        if not nom_excel:
            continue

        act, score = trouver_correspondance(nom_excel, activites_db)

        if score < SEUIL_FUZZY:
            non_trouves.append((nom_excel, act["nom"] if act else "?", round(score, 2)))
            continue

        if score < 0.95:
            scores_bas.append((nom_excel, act["nom"], round(score, 2)))

        for code in ("C1", "C2", "C3"):
            cycle_id  = cycles[code]
            a_le_cycle = row[idx[code]] == "X"
            cle = (act["id"], cycle_id)

            if a_le_cycle:
                # L'activite est prevue pour ce cycle : Gemini s'en chargera
                # On ne touche pas a la table
                continue

            if cle in deja:
                deja_presents += 1
                continue

            # Marquer comme inadapte
            inadaptes_new += 1
            if not SIMULATION:
                conn.execute(
                    "INSERT OR IGNORE INTO activite_cycle_analysee "
                    "(activite_id, cycle_id, statut) VALUES (?, ?, 'inadapte')",
                    (act["id"], cycle_id)
                )

    if not SIMULATION:
        conn.commit()

    # --- Rapport ---
    print(f"  {inadaptes_new} associations 'inadapte' {'a inserer' if SIMULATION else 'inserees'}")
    print(f"  {deja_presents} associations deja presentes en base (non touchees)")
    print()

    if scores_bas:
        print(f"  {len(scores_bas)} correspondances fuzzy (score entre {SEUIL_FUZZY} et 0.95) :")
        for excel, db, s in scores_bas[:20]:
            print(f"    [{s}] '{excel}'")
            print(f"          -> '{db}'")
        if len(scores_bas) > 20:
            print(f"    ... et {len(scores_bas) - 20} autres")
        print()

    if non_trouves:
        print(f"  AVERTISSEMENT : {len(non_trouves)} activites Excel non trouvees en base (score < {SEUIL_FUZZY}) :")
        for excel, db, s in non_trouves[:20]:
            print(f"    [{s}] '{excel}'")
            print(f"          meilleur candidat : '{db}'")
        if len(non_trouves) > 20:
            print(f"    ... et {len(non_trouves) - 20} autres")
        print()

    if SIMULATION:
        print(">>> Simulation terminee. Mettre SIMULATION = False pour appliquer.")
    else:
        print("OK Import termine.")

    conn.close()


if __name__ == "__main__":
    main()
